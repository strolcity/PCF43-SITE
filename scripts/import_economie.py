# -*- coding: utf-8 -*-
"""
import_economie.py
-------------------
Lit les fichiers xlsx du dossier data/economie/ et les charge dans economie.db.

Organisation : une fonction d'import par fichier xlsx, appelées les unes après
les autres dans main(). Chaque fonction (re)crée ses propres tables (DROP puis
CREATE) donc le script est rejouable autant de fois que nécessaire, par
exemple après mise à jour d'un fichier xlsx.

Pour ajouter un nouveau fichier plus tard : copier une des fonctions
import_xxx() ci-dessous, l'adapter, puis l'appeler dans main().
"""

import sqlite3
import openpyxl
from pathlib import Path

# ----------------------------------------------------------------------
# Configuration des chemins (à adapter si besoin)
# ----------------------------------------------------------------------
RACINE = Path(__file__).resolve().parent.parent   # dossier du site (SITE_PCF_43)
DOSSIER_ECONOMIE = RACINE / "data" / "economie"
CHEMIN_DB = RACINE / "economie.db"

FICHIER_PIB = DOSSIER_ECONOMIE / "Data_PIB.xlsx"
FICHIER_PAUVRETE = DOSSIER_ECONOMIE / "Data_Pauvrete_Patrimoine.xlsx"
FICHIER_SECU = DOSSIER_ECONOMIE / "Data_Financement_Secu.xlsx"


def lire_lignes(fichier, feuille, ligne_debut=2):
    """Lit toutes les lignes d'une feuille à partir de ligne_debut (les données,
    après les en-têtes), en ignorant les lignes vides."""
    wb = openpyxl.load_workbook(fichier, data_only=True, read_only=True)
    ws = wb[feuille]
    for row in ws.iter_rows(min_row=ligne_debut, values_only=True):
        if row[0] is None:
            continue
        yield row


# ----------------------------------------------------------------------
# 1) Data_PIB.xlsx
# ----------------------------------------------------------------------
def import_pib(cur):
    cur.execute("DROP TABLE IF EXISTS pib_niveau")
    cur.execute("""
        CREATE TABLE pib_niveau (
            trimestre TEXT PRIMARY KEY,
            annee INTEGER,
            pib_valeur_mdeur REAL,
            pib_volume_mdeur REAL,
            deflateur_indice REAL
        )
    """)
    for trimestre, annee, valeur, volume, deflateur in lire_lignes(FICHIER_PIB, "PIB_niveau"):
        cur.execute(
            "INSERT INTO pib_niveau VALUES (?,?,?,?,?)",
            (trimestre, annee, valeur, volume, deflateur),
        )

    cur.execute("DROP TABLE IF EXISTS pib_croissance")
    cur.execute("""
        CREATE TABLE pib_croissance (
            trimestre TEXT PRIMARY KEY,
            annee INTEGER,
            croissance_trim_pct REAL,
            croissance_glissante_annuelle_pct REAL
        )
    """)
    for trimestre, annee, trim_pct, glissante_pct in lire_lignes(FICHIER_PIB, "PIB_croissance"):
        cur.execute(
            "INSERT INTO pib_croissance VALUES (?,?,?,?)",
            (trimestre, annee, trim_pct, glissante_pct if glissante_pct != "" else None),
        )

    # International_comparaison est au format "large" (une colonne par pays) dans
    # le xlsx, pratique à relire à l'oeil. On la transforme ici au format "long"
    # (une ligne par pays et par année) : plus pratique pour Chart.js, notamment
    # pour le bouton "afficher/masquer un pays" prévu sur le graphique.
    cur.execute("DROP TABLE IF EXISTS pib_international")
    cur.execute("""
        CREATE TABLE pib_international (
            annee INTEGER,
            pays TEXT,
            croissance_pct REAL,
            type TEXT,
            PRIMARY KEY (annee, pays)
        )
    """)
    pays_colonnes = ["France", "Allemagne", "Etats-Unis", "Japon", "Royaume-Uni", "Zone_euro", "OCDE_moyenne"]
    for row in lire_lignes(FICHIER_PIB, "International_comparaison"):
        annee = row[0]
        if not isinstance(annee, (int, float)):
            continue  # ignore la ligne de note en bas de la feuille
        valeurs = row[1:8]
        type_donnee = row[8]
        for pays, val in zip(pays_colonnes, valeurs):
            if val is not None:
                cur.execute(
                    "INSERT INTO pib_international VALUES (?,?,?,?)",
                    (int(annee), pays, val, type_donnee),
                )


# ----------------------------------------------------------------------
# 2) Data_Pauvrete_Patrimoine.xlsx
# ----------------------------------------------------------------------
def import_pauvrete(cur):
    # La feuille Pauvrete_departements est au format "indicateur en ligne,
    # zone en colonne" (pratique à lire dans Excel). On la repivote en table
    # (zone, indicateur, valeur) : un format plus simple à requêter en SQL.
    cur.execute("DROP TABLE IF EXISTS pauvrete_filosofi")
    cur.execute("""
        CREATE TABLE pauvrete_filosofi (
            zone TEXT,
            indicateur TEXT,
            valeur REAL,
            PRIMARY KEY (zone, indicateur)
        )
    """)
    wb = openpyxl.load_workbook(FICHIER_PAUVRETE, data_only=True, read_only=True)
    ws = wb["Pauvrete_departements"]
    rows = list(ws.iter_rows(values_only=True))
    zones = rows[0][1:3]  # ("Haute-Loire (43)", "France métropolitaine")
    for row in rows[1:]:
        indicateur = row[0]
        if indicateur is None:
            continue
        for zone, valeur in zip(zones, row[1:3]):
            if valeur is not None:
                cur.execute(
                    "INSERT INTO pauvrete_filosofi VALUES (?,?,?)",
                    (zone, indicateur, valeur),
                )

    cur.execute("DROP TABLE IF EXISTS patrimoine_ifi")
    cur.execute("""
        CREATE TABLE patrimoine_ifi (
            annee INTEGER PRIMARY KEY,
            nombre_redevables INTEGER,
            patrimoine_moyen_tranche_max_keur REAL
        )
    """)
    for annee, nb, patrimoine in lire_lignes(FICHIER_PAUVRETE, "Patrimoine_IFI"):
        if not isinstance(annee, (int, float)):
            continue
        cur.execute("INSERT INTO patrimoine_ifi VALUES (?,?,?)", (int(annee), nb, patrimoine))

    cur.execute("DROP TABLE IF EXISTS ifi_deciles")
    cur.execute("""
        CREATE TABLE ifi_deciles (
            annee INTEGER,
            decile INTEGER,
            borne_inf_keur REAL,
            borne_sup_keur REAL,
            impot_net_moyen_keur REAL,
            PRIMARY KEY (annee, decile)
        )
    """)
    for annee, decile, b_inf, b_sup, impot in lire_lignes(FICHIER_PAUVRETE, "IFI_deciles_detail"):
        cur.execute(
            "INSERT INTO ifi_deciles VALUES (?,?,?,?,?)",
            (int(annee), int(decile), b_inf, b_sup, impot),
        )


# ----------------------------------------------------------------------
# 3) Data_Financement_Secu.xlsx
# ----------------------------------------------------------------------
def import_secu(cur):
    cur.execute("DROP TABLE IF EXISTS protection_sociale")
    cur.execute("""
        CREATE TABLE protection_sociale (
            annee INTEGER PRIMARY KEY,
            total_prestations_meuros REAL,
            sante_meuros REAL,
            vieillesse_survie_meuros REAL,
            famille_meuros REAL,
            emploi_meuros REAL,
            logement_meuros REAL,
            pauvrete_exclusion_meuros REAL
        )
    """)
    for row in lire_lignes(FICHIER_SECU, "Protection_sociale_par_risque"):
        cur.execute("INSERT INTO protection_sociale VALUES (?,?,?,?,?,?,?,?)", tuple(row))

    cur.execute("DROP TABLE IF EXISTS exonerations_cotisations")
    cur.execute("""
        CREATE TABLE exonerations_cotisations (
            annee INTEGER PRIMARY KEY,
            total_meuros REAL,
            allegements_generaux_meuros REAL,
            reduction_generale_bas_salaires_meuros REAL,
            mesures_contrats_particuliers_meuros REAL,
            mesures_secteurs_particuliers_meuros REAL,
            mesures_zones_particulieres_meuros REAL,
            autres_mesures_meuros REAL
        )
    """)
    for row in lire_lignes(FICHIER_SECU, "Exonerations_cotisations"):
        if row[0] is None or not isinstance(row[0], (int, float)):
            continue  # ignore la ligne de note en bas de la feuille
        cur.execute("INSERT INTO exonerations_cotisations VALUES (?,?,?,?,?,?,?,?)", tuple(row))


# ----------------------------------------------------------------------
# 4) Vues de croisement (PIB annuel + % du PIB), calculées et non stockées
#    en dur : si le PIB ou les prestations sont mis à jour, la vue se
#    recalcule automatiquement.
# ----------------------------------------------------------------------
def creer_vues(cur):
    cur.execute("DROP VIEW IF EXISTS vue_pib_annuel")
    cur.execute("""
        CREATE VIEW vue_pib_annuel AS
        SELECT annee, SUM(pib_valeur_mdeur) AS pib_valeur_mdeur_annuel
        FROM pib_niveau
        GROUP BY annee
        HAVING COUNT(*) = 4   -- ne garde que les années avec les 4 trimestres complets
    """)

    cur.execute("DROP VIEW IF EXISTS vue_protection_sociale_pct_pib")
    cur.execute("""
        CREATE VIEW vue_protection_sociale_pct_pib AS
        SELECT
            p.annee,
            p.total_prestations_meuros,
            p.vieillesse_survie_meuros,
            pa.pib_valeur_mdeur_annuel,
            100.0 * (p.total_prestations_meuros / 1000.0) / pa.pib_valeur_mdeur_annuel AS total_prestations_pct_pib,
            100.0 * (p.vieillesse_survie_meuros / 1000.0) / pa.pib_valeur_mdeur_annuel AS retraites_pct_pib
        FROM protection_sociale p
        JOIN vue_pib_annuel pa ON pa.annee = p.annee
    """)

    cur.execute("DROP VIEW IF EXISTS vue_exonerations_pct_pib")
    cur.execute("""
        CREATE VIEW vue_exonerations_pct_pib AS
        SELECT
            e.annee,
            e.total_meuros,
            e.reduction_generale_bas_salaires_meuros,
            pa.pib_valeur_mdeur_annuel,
            100.0 * (e.total_meuros / 1000.0) / pa.pib_valeur_mdeur_annuel AS total_exonerations_pct_pib
        FROM exonerations_cotisations e
        JOIN vue_pib_annuel pa ON pa.annee = e.annee
    """)


def main():
    CHEMIN_DB.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(CHEMIN_DB)
    cur = conn.cursor()

    print("Import PIB...")
    import_pib(cur)
    print("Import pauvreté / patrimoine...")
    import_pauvrete(cur)
    print("Import financement Sécu...")
    import_secu(cur)
    print("Création des vues de croisement...")
    creer_vues(cur)

    conn.commit()
    conn.close()
    print(f"Terminé. Base créée : {CHEMIN_DB}")


if __name__ == "__main__":
    main()
