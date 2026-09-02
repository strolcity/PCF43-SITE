#!/usr/bin/env python3
"""
Import du fichier immigration_origines.xlsx vers un JSON exploitable
par Chart.js (camembert 4 continents + détail pays pour le pop-up).

Usage : python3 scripts/import_origines.py data/immigration_origines.xlsx > data/origines.json
(adapter les chemins à l'arborescence réelle du repo : data/immigration/... )
"""
import sys
import json
import openpyxl

def read_continents(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        continent, effectifs, part = row[0], row[1], row[2]
        if continent is None or continent == "Ensemble":
            continue
        rows.append({"continent": continent, "effectifs": effectifs, "part_pct": part})
    return rows

def read_detail(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        continent, pays, effectifs, part = row[0], row[1], row[2], row[3]
        if continent is None:
            continue
        rows.append({"continent": continent, "pays": pays, "effectifs": effectifs, "part_pct": part})
    return rows

def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {
        "source": "INSEE, https://www.insee.fr/fr/statistiques/2861345#tableau-figure1_radio1",
        "annee": 2024,
        "champ": "France, flux d'entrées 2024",
        "immigres": {
            "continents": read_continents(wb["Continents_Immigres"]),
            "detail_pays": read_detail(wb["Detail_pays_Immigres"]),
        },
        "etrangers": {
            "continents": read_continents(wb["Continents_Etrangers"]),
            "detail_pays": read_detail(wb["Detail_pays_Etrangers"]),
        },
    }
    print(json.dumps(data, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: import_origines.py <chemin_vers_xlsx>", file=sys.stderr)
        sys.exit(1)
    main(sys.argv[1])
