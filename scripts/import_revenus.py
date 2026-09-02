#!/usr/bin/env python3
"""
Import du fichier immigration_revenus.xlsx vers un JSON exploitable
par Chart.js (décomposition des revenus 2019 + taux de pauvreté).

Usage : python3 scripts/import_revenus.py data/immigration/immigration_revenus.xlsx > data/immigration/revenus.json
"""
import sys
import json
import openpyxl

COMPOSANTES = [
    ("Revenus d'activité", 1, "rgba(255, 85, 85, 0.7)", "#ff5555"),
    ("Allocations chômage", 2, "rgba(255, 170, 0, 0.7)", "#ffaa00"),
    ("Pensions/retraites", 3, "rgba(85, 255, 85, 0.7)", "#55ff55"),
    ("Revenus patrimoine", 4, "rgba(255, 255, 85, 0.7)", "#ffff55"),
    ("Prestations sociales", 5, "rgba(85, 85, 255, 0.7)", "#5555ff"),
    ("Impôts directs", 6, "rgba(255, 85, 255, 0.7)", "#ff55ff"),
]

def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Decomposition_revenus"]
    categories = []
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
        categories.append(row[0])
        rows.append(row)

    series = []
    for label, col_idx, bg, border in COMPOSANTES:
        series.append({
            "label": label,
            "data": [r[col_idx] for r in rows],
            "backgroundColor": bg,
            "borderColor": border,
        })

    niveau_de_vie = [r[7] for r in rows]

    ws2 = wb["Taux_pauvrete"]
    pauvrete = []
    for row in ws2.iter_rows(min_row=2, max_row=3, values_only=True):
        pauvrete.append({"population": row[0], "taux_pct": row[1]})

    data = {
        "source": "INSEE, Enquête Revenus fiscaux et sociaux 2019, https://www.insee.fr/fr/statistiques/7941379?sommaire=7941491#graphique-figure3",
        "annee": 2019,
        "categories": categories,
        "series": series,
        "niveau_de_vie_moyen_eur": dict(zip(categories, niveau_de_vie)),
        "taux_pauvrete_seuil_1288e_mois": pauvrete,
    }
    print(json.dumps(data, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: import_revenus.py <chemin_vers_xlsx>", file=sys.stderr)
        sys.exit(1)
    main(sys.argv[1])
